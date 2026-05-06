"""
=============================================================================
CALCULADORA DE TAMAÑO DE MUESTRA — Basado en archivo Excel original
=============================================================================
Hojas implementadas:
  1. nEstProConN  → Estimación de proporciones CON tamaño de población conocido
  2. nEstProSinN  → Estimación de proporciones SIN tamaño de población conocido
  3. nEstMedConN  → Estimación de medias CON tamaño de población conocido
  4. nEstMedSinN  → Estimación de medias SIN tamaño de población conocido

Fórmulas utilizadas:
  - Proporciones CON N:  n = (N·Za²·p·q) / (d²·(N-1) + Za²·p·q)
  - Proporciones SIN N:  n = (Za²·p·q) / d²
  - Medias CON N:        n = (N·Za²·s²)  / (d²·(N-1) + Za²·s²)
  - Medias SIN N:        n = (Za²·s²)   / d²
=============================================================================
"""

import math
import sys
from scipy.stats import norm
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def z_alpha(nc: float) -> float:
    """Devuelve el valor Za para un nivel de confianza dado (ej. 0.95 → 1.96)."""
    return norm.ppf(1 - (1 - nc) / 2)


def validar_positivo(valor: float, nombre: str) -> None:
    if valor <= 0:
        raise ValueError(f"'{nombre}' debe ser un número positivo. Recibido: {valor}")


def validar_rango_01(valor: float, nombre: str) -> None:
    if not (0 < valor < 1):
        raise ValueError(f"'{nombre}' debe estar entre 0 y 1 (excluidos). Recibido: {valor}")


def separador(titulo: str, ancho: int = 65) -> None:
    print("\n" + "═" * ancho)
    print(f"  {titulo}")
    print("═" * ancho)


def linea(etiqueta: str, valor, unidad: str = "") -> None:
    print(f"  {etiqueta:<45} {valor}  {unidad}")


def pedir_float(mensaje: str, default: float = None) -> float:
    sufijo = f" [{default}]" if default is not None else ""
    while True:
        entrada = input(f"  {mensaje}{sufijo}: ").strip()
        if entrada == "" and default is not None:
            return default
        try:
            return float(entrada)
        except ValueError:
            print("  ⚠  Ingresa un número válido.")


def pedir_int(mensaje: str, default: int = None) -> int:
    sufijo = f" [{default}]" if default is not None else ""
    while True:
        entrada = input(f"  {mensaje}{sufijo}: ").strip()
        if entrada == "" and default is not None:
            return default
        try:
            val = int(entrada)
            return val
        except ValueError:
            print("  ⚠  Ingresa un entero válido.")


# ─────────────────────────────────────────────────────────────────────────────
# MÓDULO 1 — Proporciones CON N conocido  (hoja: nEstProConN)
# ─────────────────────────────────────────────────────────────────────────────

def hoja1_proporcion_con_N() -> dict:
    """
    Tamaño de muestra para estimación de PROPORCIONES cuando SE CONOCE N.

    Fórmula:
        n = (N · Za² · p · q) / (d² · (N - 1) + Za² · p · q)

    Parámetros del Excel original:
        NC = 0.95,  N = 1000,  Za = 1.96,  p = 0.58,  q = 0.42,  d = 0.01
    """
    separador("HOJA 1 · Proporciones — Población CONOCIDA (nEstProConN)")
    print("  Fórmula: n = (N·Za²·p·q) / (d²·(N-1) + Za²·p·q)\n")

    nc = pedir_float("Ingrese el nivel de confianza (ej. 0.95)")
    validar_rango_01(nc, "Nivel de confianza")

    N  = pedir_int("Ingrese tamaño de la población N")
    validar_positivo(N, "N")

    p  = pedir_float("ingrese la proporción esperada p (0 a 1, ej. 0.58)")
    validar_rango_01(p, "p")

    d  = pedir_float("ingrese Precisión / error máximo admisible d (ej. 0.01)")
    validar_positivo(d, "d")

    # Cálculos
    alfa = 1 - nc
    Za   = z_alpha(nc)
    q    = 1 - p
    Za2  = Za ** 2

    numerador   = N * Za2 * p * q
    denominador = (d**2) * (N - 1) + Za2 * p * q
    n_exacto    = numerador / denominador
    n_final     = math.ceil(n_exacto)

    # Mostrar resultados
    separador("RESULTADOS", 65)
    linea("Nivel de confianza (NC)", f"{nc:.4f}")
    linea("Nivel de significación α", f"{alfa:.4f}")
    linea("Valor crítico Za", f"{Za:.6f}")
    linea("Tamaño de la población N", N)
    linea("Proporción de éxito p", f"{p:.4f}")
    linea("Proporción de fracaso q = 1−p", f"{q:.4f}")
    linea("Precisión d", f"{d:.4f}")
    print()
    linea("Numerador  N·Za²·p·q", f"{numerador:.4f}")
    linea("Denominador  d²(N-1) + Za²·p·q", f"{denominador:.4f}")
    linea("n exacto", f"{n_exacto:.4f}")
    print()
    print(f"  ✔  TAMAÑO DE MUESTRA  n = {n_final}  (redondeado al entero superior)")

    return {
        "hoja": "nEstProConN",
        "NC": nc, "alfa": alfa, "Za": Za, "N": N,
        "p": p, "q": q, "d": d,
        "n_exacto": n_exacto, "n": n_final
    }


# ─────────────────────────────────────────────────────────────────────────────
# MÓDULO 2 — Proporciones SIN N conocido  (hoja: nEstProSinN)
# ─────────────────────────────────────────────────────────────────────────────

def hoja2_proporcion_sin_N() -> dict:
    """
    Tamaño de muestra para estimación de PROPORCIONES cuando N es DESCONOCIDO.

    Fórmula:
        n = (Za² · p · q) / d²

    Parámetros del Excel original:
        NC = 0.95,  Za = 1.96,  p = 0.50,  q = 0.50,  d = 0.01
        n = 9604
    """
    separador("HOJA 2 · Proporciones — Población DESCONOCIDA (nEstProSinN)")
    print("  Fórmula: n = (Za²·p·q) / d²\n")

    nc = pedir_float("ingrese nivel de confianza (ej. 0.95)")
    validar_rango_01(nc, "Nivel de confianza")

    p  = pedir_float("ingrese proporción esperada p (0 a 1). Si es desconocida usa 0.5")
    validar_rango_01(p, "p")

    d  = pedir_float("ingrese precisión / error máximo admisible d (ej. 0.01)")
    validar_positivo(d, "d")

    # Cálculos
    alfa = 1 - nc
    Za   = z_alpha(nc)
    q    = 1 - p
    Za2  = Za ** 2

    n_exacto = (Za2 * p * q) / (d**2)
    n_final  = math.ceil(n_exacto)

    # Cálculo adicional con p=0.10 (como menciona el Excel en la nota)
    p_alt   = 0.10
    q_alt   = 1 - p_alt
    n_alt   = math.ceil((Za2 * p_alt * q_alt) / (d**2))

    # Mostrar resultados
    separador("RESULTADOS", 65)
    linea("Nivel de confianza (NC)", f"{nc:.4f}")
    linea("Nivel de significación α", f"{alfa:.4f}")
    linea("Valor crítico Za", f"{Za:.6f}")
    linea("Proporción de éxito p", f"{p:.4f}")
    linea("Proporción de fracaso q = 1−p", f"{q:.4f}")
    linea("Precisión d", f"{d:.4f}")
    print()
    linea("n exacto (Za²·p·q / d²)", f"{n_exacto:.4f}")
    print()
    print(f"  ✔  TAMAÑO DE MUESTRA  n = {n_final}  (proporción p = {p})")
    print(f"  ✔  TAMAÑO DE MUESTRA  n = {n_alt}  (proporción alternativa p = {p_alt})")
    print()
    print("  NOTA del Excel: Si p es desconocida se usa p=0.50 (maximiza la muestra).")

    return {
        "hoja": "nEstProSinN",
        "NC": nc, "alfa": alfa, "Za": Za,
        "p": p, "q": q, "d": d,
        "n_exacto": n_exacto, "n": n_final,
        "n_p50": n_final, "n_p10": n_alt
    }


# ─────────────────────────────────────────────────────────────────────────────
# MÓDULO 3 — Medias CON N conocido  (hoja: nEstMedConN)
# ─────────────────────────────────────────────────────────────────────────────

def hoja3_media_con_N() -> dict:
    """
    Tamaño de muestra para estimación de MEDIAS cuando SE CONOCE N.

    Fórmula:
        n = (N · Za² · s²) / (d² · (N - 1) + Za² · s²)

    Parámetros del Excel original:
        NC = 0.95,  N = 15000,  Za = 1.96,  s = 0.50,  d = 0.01
        n = 5856
    """
    separador("HOJA 3 · Medias — Población CONOCIDA (nEstMedConN)")
    print("  Fórmula: n = (N·Za²·s²) / (d²·(N-1) + Za²·s²)\n")

    nc = pedir_float("ingrese nivel de confianza (ej. 0.95)")
    validar_rango_01(nc, "Nivel de confianza")

    N  = pedir_int("ingrese tamaño de la población N")
    validar_positivo(N, "N")

    s  = pedir_float("ingrese desviación estándar s (ej. 0.50)")
    validar_positivo(s, "s")

    d  = pedir_float("Ingrese precisión / error máximo admisible d (ej. 0.01)")
    validar_positivo(d, "d")

    # Cálculos
    alfa = 1 - nc
    Za   = z_alpha(nc)
    Za2  = Za ** 2
    s2   = s ** 2

    numerador   = N * Za2 * s2
    denominador = (d**2) * (N - 1) + Za2 * s2
    n_exacto    = numerador / denominador
    n_final     = math.ceil(n_exacto)

    # Mostrar resultados
    separador("RESULTADOS", 65)
    linea("Nivel de confianza (NC)", f"{nc:.4f}")
    linea("Nivel de significación α", f"{alfa:.4f}")
    linea("Valor crítico Za", f"{Za:.6f}")
    linea("Tamaño de la población N", N)
    linea("Desviación estándar s", f"{s:.4f}")
    linea("Varianza s²", f"{s2:.4f}")
    linea("Precisión d", f"{d:.4f}")
    print()
    linea("Numerador  N·Za²·s²", f"{numerador:.4f}")
    linea("Denominador  d²(N-1) + Za²·s²", f"{denominador:.4f}")
    linea("n exacto", f"{n_exacto:.4f}")
    print()
    print(f"  ✔  TAMAÑO DE MUESTRA  n = {n_final}  (redondeado al entero superior)")

    return {
        "hoja": "nEstMedConN",
        "NC": nc, "alfa": alfa, "Za": Za, "N": N,
        "s": s, "s2": s2, "d": d,
        "n_exacto": n_exacto, "n": n_final
    }


# ─────────────────────────────────────────────────────────────────────────────
# MÓDULO 4 — Medias SIN N conocido  (hoja: nEstMedSinN)
# ─────────────────────────────────────────────────────────────────────────────

def hoja4_media_sin_N() -> dict:
    """
    Tamaño de muestra para estimación de MEDIAS cuando N es DESCONOCIDO.

    Fórmula:
        n = (Za² · s²) / d²

    Parámetros del Excel original:
        NC = 0.95,  Za = 1.96,  s = 0.50,  d = 0.01
        n = 9604
    """
    separador("HOJA 4 · Medias — Población DESCONOCIDA (nEstMedSinN)")
    print("  Fórmula: n = (Za²·s²) / d²\n")

    nc = pedir_float("ingrese nivel de confianza (ej. 0.95)")
    validar_rango_01(nc, "Nivel de confianza")

    s  = pedir_float("ingrese Desviación estándar s (ej. 0.50)")
    validar_positivo(s, "s")

    d  = pedir_float("ingrese Precisión / error máximo admisible d (ej. 0.01)")
    validar_positivo(d, "d")

    # Cálculos
    alfa = 1 - nc
    Za   = z_alpha(nc)
    Za2  = Za ** 2
    s2   = s ** 2

    n_exacto = (Za2 * s2) / (d**2)
    n_final  = math.ceil(n_exacto)

    # Mostrar resultados
    separador("RESULTADOS", 65)
    linea("Nivel de confianza (NC)", f"{nc:.4f}")
    linea("Nivel de significación α", f"{alfa:.4f}")
    linea("Valor crítico Za", f"{Za:.6f}")
    linea("Desviación estándar s", f"{s:.4f}")
    linea("Varianza s²", f"{s2:.4f}")
    linea("Precisión d", f"{d:.4f}")
    print()
    linea("n exacto (Za²·s² / d²)", f"{n_exacto:.4f}")
    print()
    print(f"  ✔  TAMAÑO DE MUESTRA  n = {n_final}  (redondeado al entero superior)")

    return {
        "hoja": "nEstMedSinN",
        "NC": nc, "alfa": alfa, "Za": Za,
        "s": s, "s2": s2, "d": d,
        "n_exacto": n_exacto, "n": n_final
    }


# ─────────────────────────────────────────────────────────────────────────────
# MÓDULO 5 — Estadísticas descriptivas adicionales (BONUS)
# ─────────────────────────────────────────────────────────────────────────────

def hoja5_estadisticas_descriptivas() -> dict:
    """
    Cálculo opcional de estadísticas descriptivas sobre una muestra:
      - Frecuencias absolutas y relativas
      - Media, mediana, moda
      - Varianza, desviación estándar
      - Cuartiles (Q1, Q2, Q3), deciles, percentiles
      - Intervalo de confianza para la media
      - Error estándar
    """
    separador("HOJA 5 · Estadísticas Descriptivas (BONUS)")
    print("  Ingresa los datos de tu muestra separados por comas (ej. 10,12,15,9,11):")
    while True:
        raw = input("  Datos: ").strip()
        try:
            datos = [float(x.strip()) for x in raw.split(",") if x.strip()]
            if len(datos) < 2:
                print("  ⚠  Ingresa al menos 2 valores.")
                continue
            break
        except ValueError:
            print("  ⚠  Solo se aceptan números separados por comas.")

    nc = pedir_float("Nivel de confianza para el IC (ej. 0.95)", 0.95)
    validar_rango_01(nc, "Nivel de confianza")

    s_datos = pd.Series(datos)
    n       = len(s_datos)
    media   = s_datos.mean()
    mediana = s_datos.median()
    moda_vals = s_datos.mode().tolist()
    varianza  = s_datos.var(ddof=1)
    desvest   = s_datos.std(ddof=1)
    error_est = desvest / math.sqrt(n)
    Za        = z_alpha(nc)
    ic_inf    = media - Za * error_est
    ic_sup    = media + Za * error_est

    # Cuartiles, deciles, percentiles seleccionados
    q1, q2, q3 = s_datos.quantile([0.25, 0.50, 0.75]).values
    deciles     = {f"D{i}": s_datos.quantile(i / 10) for i in range(1, 10)}
    percentiles = {f"P{i}": s_datos.quantile(i / 100) for i in [5, 10, 25, 50, 75, 90, 95]}

    # Frecuencias
    freq = s_datos.value_counts().sort_index()
    freq_rel = freq / n

    separador("RESULTADOS", 65)
    linea("n (tamaño de muestra)", n)
    linea("Media", f"{media:.6f}")
    linea("Mediana", f"{mediana:.6f}")
    linea("Moda", ", ".join(str(m) for m in moda_vals))
    linea("Varianza (s²)", f"{varianza:.6f}")
    linea("Desviación estándar (s)", f"{desvest:.6f}")
    linea("Error estándar (s/√n)", f"{error_est:.6f}")
    print()
    linea(f"IC {nc*100:.0f}% para la media  [{ic_inf:.4f}", f"{ic_sup:.4f}]")
    print()
    linea("Q1 (percentil 25)", f"{q1:.4f}")
    linea("Q2 (percentil 50 = Mediana)", f"{q2:.4f}")
    linea("Q3 (percentil 75)", f"{q3:.4f}")
    linea("RIQ (Q3 - Q1)", f"{q3 - q1:.4f}")
    print()
    print("  DECILES:")
    for k, v in deciles.items():
        print(f"    {k}: {v:.4f}")
    print()
    print("  PERCENTILES CLAVE:")
    for k, v in percentiles.items():
        print(f"    {k}: {v:.4f}")
    print()
    print("  FRECUENCIAS ABSOLUTAS Y RELATIVAS:")
    print(f"  {'Valor':>10}  {'Frec. Abs.':>12}  {'Frec. Rel.':>12}")
    print("  " + "-" * 38)
    for val, f_abs in freq.items():
        f_rel = freq_rel[val]
        print(f"  {val:>10}  {f_abs:>12}  {f_rel:>12.4f}")

    return {
        "hoja": "EstDesc",
        "n": n, "media": media, "mediana": mediana, "moda": moda_vals,
        "varianza": varianza, "desvest": desvest, "error_est": error_est,
        "NC": nc, "ic_inf": ic_inf, "ic_sup": ic_sup,
        "Q1": q1, "Q2": q2, "Q3": q3,
        "deciles": deciles, "percentiles": percentiles,
        "frecuencias": {str(k): int(v) for k, v in freq.items()},
        "frecuencias_rel": {str(k): round(float(v), 6) for k, v in freq_rel.items()},
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR A EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _estilo_encabezado(ws, fila: int, col_ini: int, col_fin: int,
                        texto: str, color: str = "1F4E79") -> None:
    ws.merge_cells(start_row=fila, start_column=col_ini,
                   end_row=fila, end_column=col_fin)
    c = ws.cell(row=fila, column=col_ini, value=texto)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", start_color=color)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _fila_dato(ws, fila: int, etiqueta: str, valor, formula: str = "") -> None:
    ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=False)
    ws.cell(row=fila, column=2, value=valor)
    if formula:
        ws.cell(row=fila, column=3, value=formula).font = Font(italic=True, color="555555")


def exportar_excel(resultados: list, ruta: str = "resultados_muestras.xlsx") -> None:
    """Exporta todos los resultados a un archivo Excel con hojas separadas y formato profesional."""
    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for res in resultados:
        hoja = res["hoja"]
        ws   = wb.create_sheet(title=hoja)
        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 40

        # Cabecera
        titulos = {
            "nEstProConN":  "Proporciones — Población CONOCIDA",
            "nEstProSinN":  "Proporciones — Población DESCONOCIDA",
            "nEstMedConN":  "Medias — Población CONOCIDA",
            "nEstMedSinN":  "Medias — Población DESCONOCIDA",
            "EstDesc":      "Estadísticas Descriptivas",
        }
        _estilo_encabezado(ws, 1, 1, 3, titulos.get(hoja, hoja))

        if hoja in ("nEstProConN", "nEstProSinN", "nEstMedConN", "nEstMedSinN"):
            _estilo_encabezado(ws, 2, 1, 3, "Parámetros de Entrada", "2E75B6")
            filas_param = []

            if hoja == "nEstProConN":
                filas_param = [
                    ("Nivel de confianza NC",         res["NC"],   ""),
                    ("Nivel de significación α",       res["alfa"], "= 1 − NC"),
                    ("Valor crítico Za",               round(res["Za"], 6), "= NORM.INV(1−α/2, 0, 1)"),
                    ("Tamaño de la población N",       res["N"],   ""),
                    ("Proporción de éxito p",          res["p"],    ""),
                    ("Proporción de fracaso q",        res["q"],    "= 1 − p"),
                    ("Precisión d",                    res["d"],    ""),
                ]
                formula_txt = "n = (N·Za²·p·q) / (d²·(N−1) + Za²·p·q)"

            elif hoja == "nEstProSinN":
                filas_param = [
                    ("Nivel de confianza NC",         res["NC"],   ""),
                    ("Nivel de significación α",       res["alfa"], "= 1 − NC"),
                    ("Valor crítico Za",               round(res["Za"], 6), "= NORM.INV(1−α/2, 0, 1)"),
                    ("Proporción de éxito p",          res["p"],   ""),
                    ("Proporción de fracaso q",        res["q"],   "= 1 − p"),
                    ("Precisión d",                    res["d"],   ""),
                ]
                formula_txt = "n = (Za²·p·q) / d²"

            elif hoja == "nEstMedConN":
                filas_param = [
                    ("Nivel de confianza NC",         res["NC"],   ""),
                    ("Nivel de significación α",       res["alfa"], "= 1 − NC"),
                    ("Valor crítico Za",               round(res["Za"], 6), "= NORM.INV(1−α/2, 0, 1)"),
                    ("Tamaño de la población N",       res["N"],   ""),
                    ("Desviación estándar s",          res["s"],    ""),
                    ("Varianza s²",                    res["s2"],   "= s²"),
                    ("Precisión d",                    res["d"],    ""),
                ]
                formula_txt = "n = (N·Za²·s²) / (d²·(N−1) + Za²·s²)"

            elif hoja == "nEstMedSinN":
                filas_param = [
                    ("Nivel de confianza NC",         res["NC"],   ""),
                    ("Nivel de significación α",       res["alfa"], "= 1 − NC"),
                    ("Valor crítico Za",               round(res["Za"], 6), "= NORM.INV(1−α/2, 0, 1)"),
                    ("Desviación estándar s",          res["s"],   ""),
                    ("Varianza s²",                    res["s2"],   "= s²"),
                    ("Precisión d",                    res["d"],   ""),
                ]
                formula_txt = "n = (Za²·s²) / d²"

            for i, (etq, val, form) in enumerate(filas_param, start=3):
                _fila_dato(ws, i, etq, val, form)
                for col in range(1, 4):
                    ws.cell(row=i, column=col).border = border

            r = 3 + len(filas_param)
            _estilo_encabezado(ws, r, 1, 3, "Resultados", "375623")
            _fila_dato(ws, r+1, "n exacto (sin redondear)",
                       round(res["n_exacto"], 6), formula_txt)
            c_n = ws.cell(row=r+2, column=1, value="✔  TAMAÑO DE MUESTRA n (redondeado)")
            c_n.font = Font(bold=True, size=12)
            ws.cell(row=r+2, column=2, value=res["n"]).font = Font(bold=True, size=12, color="375623")
            for c in range(1, 4):
                ws.cell(row=r+1, column=c).border = border
                ws.cell(row=r+2, column=c).border = border

            # Notas adicionales para nEstProSinN
            if hoja == "nEstProSinN" and "n_p10" in res:
                ws.cell(row=r+4, column=1,
                        value="n con proporción alternativa p=0.10").font = Font(italic=True)
                ws.cell(row=r+4, column=2, value=res["n_p10"])

        elif hoja == "EstDesc":
            fila = 2
            secciones = [
                ("Tamaño n",           res["n"],           ""),
                ("Media",              round(res["media"],6),"= PROMEDIO(datos)"),
                ("Mediana",            round(res["mediana"],6),"= MEDIANA(datos)"),
                ("Moda",               ", ".join(str(m) for m in res["moda"]), "= MODA(datos)"),
                ("Varianza (s²)",      round(res["varianza"],6),"= VAR(datos)"),
                ("Desviación estándar (s)", round(res["desvest"],6),"= DESVEST(datos)"),
                ("Error estándar (s/√n)",   round(res["error_est"],6),"= s / RAIZ(n)"),
                (f"IC {res['NC']*100:.0f}% inferior", round(res["ic_inf"],6),"= media − Za·EE"),
                (f"IC {res['NC']*100:.0f}% superior", round(res["ic_sup"],6),"= media + Za·EE"),
                ("Q1", round(res["Q1"],6),"= PERCENTIL(datos,0.25)"),
                ("Q2 (Mediana)", round(res["Q2"],6),"= PERCENTIL(datos,0.50)"),
                ("Q3", round(res["Q3"],6),"= PERCENTIL(datos,0.75)"),
                ("RIQ (Q3−Q1)", round(res["Q3"]-res["Q1"],6),""),
            ]
            for etq, val, form in secciones:
                _fila_dato(ws, fila, etq, val, form)
                for c in range(1, 4):
                    ws.cell(row=fila, column=c).border = border
                fila += 1

            fila += 1
            _estilo_encabezado(ws, fila, 1, 3, "Deciles", "2E75B6")
            fila += 1
            for k, v in res["deciles"].items():
                _fila_dato(ws, fila, k, round(v, 6))
                fila += 1

            fila += 1
            _estilo_encabezado(ws, fila, 1, 3, "Percentiles clave", "2E75B6")
            fila += 1
            for k, v in res["percentiles"].items():
                _fila_dato(ws, fila, k, round(v, 6))
                fila += 1

            fila += 1
            _estilo_encabezado(ws, fila, 1, 3, "Frecuencias", "2E75B6")
            fila += 1
            ws.cell(row=fila, column=1, value="Valor").font = Font(bold=True)
            ws.cell(row=fila, column=2, value="Frec. Absoluta").font = Font(bold=True)
            ws.cell(row=fila, column=3, value="Frec. Relativa").font = Font(bold=True)
            fila += 1
            for k in sorted(res["frecuencias"].keys()):
                ws.cell(row=fila, column=1, value=float(k))
                ws.cell(row=fila, column=2, value=res["frecuencias"][k])
                ws.cell(row=fila, column=3, value=res["frecuencias_rel"].get(k, ""))
                fila += 1

    wb.save(ruta)
    print(f"\n  📄 Resultados exportados a: {ruta}")


# ─────────────────────────────────────────────────────────────────────────────
# MENÚ PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

MENU = """
╔══════════════════════════════════════════════════════════════════╗
║        CALCULADORA DE TAMAÑO DE MUESTRA — Python / Excel        ║
╠══════════════════════════════════════════════════════════════════╣
║  1. Proporciones   — Población CONOCIDA      (nEstProConN)       ║
║  2. Proporciones   — Población DESCONOCIDA   (nEstProSinN)       ║
║  3. Medias         — Población CONOCIDA      (nEstMedConN)       ║
║  4. Medias         — Población DESCONOCIDA   (nEstMedSinN)       ║
║  5. Estadísticas Descriptivas (BONUS)                            ║
╠══════════════════════════════════════════════════════════════════╣
║  E. Exportar todos los resultados a Excel                        ║
║  0. Salir                                                        ║
╚══════════════════════════════════════════════════════════════════╝
"""

FUNCIONES = {
    "1": hoja1_proporcion_con_N,
    "2": hoja2_proporcion_sin_N,
    "3": hoja3_media_con_N,
    "4": hoja4_media_sin_N,
    "5": hoja5_estadisticas_descriptivas,
}


def main() -> None:
    resultados = []

    while True:
        print(MENU)
        opcion = input("  Elige una opción: ").strip().upper()

        if opcion == "0":
            print("\n  Hasta luego. ¡Que tus muestras sean representativas! 👋\n")
            sys.exit(0)

        elif opcion in FUNCIONES:
            try:
                res = FUNCIONES[opcion]()
                resultados.append(res)
                input("\n  Presiona Enter para continuar...")
            except ValueError as e:
                print(f"\n  ⚠  Error de validación: {e}")
                input("  Presiona Enter para continuar...")
            except KeyboardInterrupt:
                print("\n  Operación cancelada.")

        elif opcion == "E":
            if not resultados:
                print("\n  ⚠  No hay resultados que exportar. Ejecuta al menos un cálculo.")
                input("  Presiona Enter para continuar...")
            else:
                ruta = input("  Nombre del archivo Excel [resultados_muestras.xlsx]: ").strip()
                ruta = ruta if ruta else "resultados_muestras.xlsx"
                if not ruta.endswith(".xlsx"):
                    ruta += ".xlsx"
                try:
                    exportar_excel(resultados, ruta)
                except Exception as e:
                    print(f"\n  ⚠  Error al exportar: {e}")
                input("  Presiona Enter para continuar...")

        else:
            print("\n  ⚠  Opción no válida. Intenta de nuevo.")


if __name__ == "__main__":
    main()
    